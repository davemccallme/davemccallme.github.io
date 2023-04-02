# -*- coding: utf-8 -*-
"""
Created on Tue May  7 19:47:27 2019

@author: DavidM
"""

### Automate the Boring Stuff [ATBS] 43 Editing Excel Spreadsheets  ###

import openpyxl
import os

wb = openpyxl.Workbook()                    # creates a NEW workbook
wb.get_sheet_names()

sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 42                            # changes the cells value
sheet['A2'] = 'Hello'

os.chdir('C:\Anaconda3\Scripts\MyPythonScripts\AutomateTheBoringStuff')
wb.save('ATBS43_pyexample.xlsx')
sheet2 = wb.create_sheet()
wb.get_sheet_names()

                                            # wb.create_sheet(index=0, title='My Other Sheet')