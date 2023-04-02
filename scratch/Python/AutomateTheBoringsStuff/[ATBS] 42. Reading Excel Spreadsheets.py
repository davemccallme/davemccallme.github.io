# -*- coding: utf-8 -*-
"""
Created on Tue May  7 19:09:53 2019

@author: DavidM
"""

# Automate the Boring Stuff [ATBS] 42 Reading Excel Spreadsheets 
# Ref: ATBS 42 Reading Excel Spreadsheets.xlsx

import openpyxl, os
os.chdir('C:\Anaconda3\Scripts\MyPythonScripts\AutomateTheBoringStuff')

workbook = openpyxl.load_workbook('ATBS 42 Reading Excel Spreadsheets.xlsx') #open the excel file
type(workbook)
sheet = workbook.get_sheet_by_name('Sheet1')  #returns sheet object

workbook.get_sheet_names()  #when you don'tknow names of workbook, returns list of string values of sheet names

cell = sheet['A1']  #returns "cell" (a cell in excel) object
cell.value #retuns the value contained in the cell

#sheet['C1'].value

#str(sheet['C1'].value)
#sheet.cell(row=1, column=2)
for i in range(1,8):
        print(i,sheet.cell(row=i, column=2).value)  #retruns list of rows i in col 2
        
        
